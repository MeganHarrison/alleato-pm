#!/usr/bin/env python3
"""Quick preview of Vermillion budget Excel file."""

import openpyxl
import sys

file_path = "/Users/meganharrison/Documents/github/alleato-procore/docs/vermillian-budget.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

print(f"\n{'='*120}")
print(f"VERMILLION RISE BUDGET - Complete Preview ({sheet.max_row - 1} line items)")
print(f"{'='*120}\n")

# Get headers
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

# Print header
print(f"{'#':<4} {'Budget Code':<45} {'Original':<12} {'Revised':<12} {'Forecast':<12} {'Over/Under':<12}")
print("-" * 120)

total_original = 0
total_revised = 0
total_forecast = 0

# Print all rows
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
    if not any(row):  # Skip empty rows
        continue

    budget_code = row[0] if row[0] else "N/A"
    original = row[1] if len(row) > 1 and row[1] is not None else 0
    revised = row[6] if len(row) > 6 and row[6] is not None else 0
    forecast = row[12] if len(row) > 12 and row[12] is not None else 0
    over_under = row[13] if len(row) > 13 and row[13] is not None else 0

    total_original += original
    total_revised += revised
    total_forecast += forecast

    print(f"{i:<4} {str(budget_code):<45} ${original:>10,.0f} ${revised:>10,.0f} ${forecast:>10,.0f} ${over_under:>10,.0f}")

print("-" * 120)
print(f"{'TOTALS':<50} ${total_original:>10,.0f} ${total_revised:>10,.0f} ${total_forecast:>10,.0f}")
print(f"{'='*120}\n")

print("Field Mapping for Import:")
print(f"  Budget Code → cost_code_id (will create cost codes)")
print(f"  Original Budget → original_budget_amount")
print(f"  Budget Modifications → budget_modifications")
print(f"  Revised Budget → revised_budget")
print(f"  Forecast to Complete → forecast_to_complete")
print(f"  And other financial tracking fields...\n")
